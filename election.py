import re
import sys
import chardet
from openpyxl import Workbook

# REG EX - VS CODE = ^Name\s
# Check multiple cases in Input txt -> 
# From - Name : / Husband's Name: / Father's Name:
# To   - Name: / Husband's Name : / Father's Name :

def extract_info(file_path):
    
    with open(file_path, 'rb') as f:
        result = chardet.detect(f.read())
    encoding = result['encoding']

    # Open the input file with the detected encoding
    with open(file_path, 'r', encoding=encoding) as f:
        data = f.read()
        
    #with open(file_path, 'r') as f:
        #data = f.read()

    # Define regular expressions to extract Name, Age and Gender from input data
    name_regex = r'Name:\s*([^\n]+)'    
    age_regex = r'Age\s*:\s*(\d+)'
    gender_regex = r'Gender\s*:\s*([A-Za-z]+)'

    # Find all occurrences of the regex in the input data
    name_matches = re.findall(name_regex, data)
    age_matches = re.findall(age_regex, data)
    gender_matches = re.findall(gender_regex, data)

    # Create a list of dictionaries to store the extracted information
    results = []
    for i in range(len(name_matches)):
        result = {
            'Name': name_matches[i].strip(),
            'Age': age_matches[i].strip() if age_matches and len(age_matches) > i and age_matches[i] else None,
            'Gender': gender_matches[i].strip().lower() if gender_matches and len(age_matches) > i and gender_matches[i] else None,
        }
        results.append(result)

    return results

if len(sys.argv) != 2:
    print("Missing File Name")
    exit()

# Test the function
results = extract_info('.\\InputData\\'+sys.argv[1])

# Create a new Excel workbook and worksheet to store the results
workbook = Workbook()
worksheet = workbook.active
worksheet.title = 'Results'

# Write the header row
worksheet.cell(row=1, column=1, value='Name')
worksheet.cell(row=1, column=2, value='Age')
worksheet.cell(row=1, column=3, value='Gender')

# Write the extracted information to the worksheet
for i, result in enumerate(results):
    row = i + 2
    worksheet.cell(row=row, column=1, value=result['Name'])
    worksheet.cell(row=row, column=2, value=result['Age'])
    worksheet.cell(row=row, column=3, value=result['Gender'])

# Save the workbook
workbook.save('.\\OutputData\\'+sys.argv[1].replace('.txt', '.xlsx'))
